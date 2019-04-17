# encoding = 'utf-8'
import importlib, inspect, json, os, sys, time, traceback, urllib, random, platform, yaml
from http.client import HTTPResponse
from threading import Thread
from datetime import datetime

from haf.common.sigleton import SingletonType
from openpyxl import load_workbook
from haf.apihelper import Request, Response
from haf.case import BaseCase
from haf.common.database import MysqlTool, SqlServerTool, SQLConfig
from haf.common.exception import FailLoadCaseFromPyException
from haf.common.log import Log
from haf.common.httprequest import HttpController
from haf.config import *
from haf.mark import test, skip

logger = Log.getLogger(__name__)


class Utils(object):

    @staticmethod
    def sql_execute(sqlconfig: SQLConfig, sqlscript: str, **kwargs)->list:
        '''
        sql_execute : connect to sqlconfig & execute sqlscript
        :param sqlconfig: SQLConfig
        :param sqlscript: sql script
        :param kwargs: others params
        :return: sql execute data
        '''
        key = kwargs.get("key")
        func_obj = None
        data = None
        try:
            sqlconfig = sqlconfig
            if sqlconfig.protocol == "mysql":
                 func_obj = MysqlTool()
            elif sqlconfig.protocol == "sqlserver":
                 func_obj = SqlServerTool()
            # elif sqlconfig.protocol == "redis":
            #     func_obj = RedisTool()
            # elif sqlconfig.protocol == "neo4j":
            #     func_obj = Neo4j()

            data = func_obj.connect_execute(sqlconfig, sqlscript, **kwargs)
            return data
        except Exception as e:
            logger.info(f"{key} {e}", __name__)
        finally:
            logger.info(f"{key} {func_obj}", __name__)
            if func_obj is not None:
                func_obj.close()

    @staticmethod
    def get_rows_from_xlsx(filename: str)->dict:
        '''
        get_rows_from_xlsx : read rows from xlsx file (filename)
        :param filename: xlsx filename
        :return: dict
        '''
        if not filename.endswith("xlsx"):
            return {}
        if not os.path.exists(filename):
            logger.error("{} not fount file : {}".format("system$%util$%", filename), __name__)
            raise FileNotFoundError
        try:
            header = []
            config_header = []
            dbconfig_header = []
            data = []
            config_data = []
            dbconfig_data = []

            result = {}
            result["testcases"] = []
            result["dbconfig"] = []
            result["config"] = []

            xlsx = load_workbook(filename)
            sheet_names = xlsx.sheetnames
            if "testcases" not in sheet_names or "config" not in sheet_names:
                logger.log(f"not fount sheet in {filename}", __name__)
                return {}
            testcases = xlsx["testcases"].rows
            config = xlsx["config"].rows
            dbconfig = xlsx["dbconfig"].rows

            for row in testcases:
                header = [cell.value for cell in row if cell.value is not None]
                break

            for row in testcases:
                data.append([cell.value for cell in row])

            for row in config:
                config_header = [cell.value for cell in row if cell.value is not None]
                break

            for row in config:
                config_data.append([cell.value for cell in row])

            for row in dbconfig:
                dbconfig_header = [cell.value for cell in row if cell.value is not None]
                break

            for row in dbconfig:
                dbconfig_data.append([cell.value for cell in row])

            for d in data:
                result["testcases"].append(dict(zip(header, d)))

            for d in config_data:
                result["config"].append(dict(zip(config_header, d)))

            for d in dbconfig_data:
                result["dbconfig"].append(dict(zip(dbconfig_header, d)))
            return result
        except Exception as e:
            logger.error(f"{filename} {e} ", __name__)

    @staticmethod
    def load_from_json(file_name: str)->dict:
        '''
        load_from_json : load json obj from file_name
        :param file_name:
        :return:
        '''
        try:
            if os.path.exists(file_name):
                f = open(file_name, 'r', encoding='utf-8')
                result = json.loads(f.read())
                f.close()
                return result
            else:
                raise FileNotFoundError
        except Exception as e:
            logger.error(f"{file_name} {e} ", __name__)

    @staticmethod
    def load_from_yml(file_name: str)-> dict:
        '''
        load_from_yml : load dict from yml file (file_name)
        :param file_name:
        :return:
        '''
        try:
            if os.path.exists(file_name):
                f = open(file_name, 'r', encoding='utf-8')
                result = yaml.load(f)
                f.close()
                return result
            else:
                raise FileNotFoundError
        except Exception as e:
            logger.error(f"{file_name} {e} ", __name__)

    @staticmethod
    def get_path(path):
        path, file = os.path.split(path)
        return path, file

    @staticmethod
    def get_class_from_py(module: str)->list:
        '''
        get_class_from_py : get classes from py file (module)
        :param module: str, module file name
        :return: list [classes]
        '''
        built_in_list = ['__class__', '__delattr__', '__dict__', '__dir__', '__doc__', '__eq__', '__format__', '__ge__', '__getattribute__', '__gt__', '__hash__', '__init__', '__init_subclass__', '__le__', '__lt__', '__module__', '__ne__', '__new__', '__reduce__', '__reduce_ex__', '__repr__', '__setattr__', '__sizeof__', '__str__', '__subclasshook__', '__weakref__']
        attr_list = ['AttrNoneList', 'bench_name', 'error_msg', 'expect', 'id', 'mark', 'name', 'run', 'subid', 'type']
        attr_none_list = ["BaseCase", "PyCase", "HttpApiCase"]
        class_list = []
        for attr in dir(module):
            class_attr = getattr(module, attr)
            class_temp = {attr: class_attr}
            if attr not in built_in_list and inspect.isclass(class_attr) and attr not in attr_none_list:
                if issubclass(class_attr, BaseCase) and isinstance(class_attr(), BaseCase):
                    class_list.append(class_temp)
        return class_list if len(class_list) > 0 else None

    @staticmethod
    def get_case_from_class(class_list: list)->dict:
        '''
        get_case_from_class : get case from class of class_list
        :param class_list: list
        :return: dict
        '''
        case_dict = {}
        for cl in class_list.keys():
            case_dict[cl] = {}
            case_dict[cl]["classes"] = []
            suites = class_list.get(cl)
            if suites is None:
                raise FailLoadCaseFromPyException
            for suite in suites:
                suite_dict = {}
                for class_temp_key in suite.keys():
                    suite_dict[class_temp_key] = {"cases": [], "request": Request()}
                    class_temp = suite.get(class_temp_key)
                    for func_key in dir(class_temp):
                        func = getattr(class_temp, func_key)
                        func_temp = {}
                        if isinstance(func, test):
                            func_temp[func_key] = func
                            suite_dict[class_temp_key]["cases"].append(func_temp)
                        elif isinstance(func, skip):
                            func_temp[func_key] = func
                            suite_dict[class_temp_key]["cases"].append(func_temp)
                        elif inspect.isfunction(func):
                            if hasattr(func, "mark") and func.mark == "test":
                                func_temp[func_key] = func
                                suite_dict[class_temp_key]["cases"].append(func_temp)
                        elif isinstance(func, Request):
                            suite_dict[class_temp_key]["request"] = func
                case_dict[cl]["classes"].append(suite_dict)
        return case_dict

    @staticmethod
    def load_dict_to_case(case_dict: dict):
        '''
        load_dict_to_case : load case from dict
        :param case_dict:
        :return:
        '''
        cases = []
        id = 0
        name = ""
        for suites_key in case_dict.keys():
            suites = case_dict.get(suites_key)
            for suite in suites.get("classes"):
                for suite_key in suite.keys():
                    suite_temp = suite.get(suite_key)
                    id += 1
                    subid = 1
                    for case in suite_temp.get("cases"):

                        for key in case.keys():
                            case_temp = {"id": id, "subid": subid, "name": key, "run": case.get(key).run if hasattr(case.get(key), 'run') else False, "reason": case.get(key).reason if hasattr(case.get(key), 'reason') else None}
                            case_temp["func"] = key
                            case_temp["suite"] = suite_key
                            case_temp["request"] = suite_temp.get("request", Request())
                            func = case.get(key)
                            if hasattr(func, "params"):
                                for param in func.params:
                                    case_temp_1 = case_temp.copy()
                                    case_temp_1["subid"] = subid
                                    case_temp_1["param"] = param
                                    cases.append(case_temp_1)
                                    subid += 1
                            else:
                                subid += 1
                                cases.append(case_temp)
        return cases

    @staticmethod
    def load_from_py(file_name: str)->dict:
        '''
        load_from_py : load from py
        :param file_name: str, filename of python
        :return: dict
        '''
        try:
            py_config = {}
            py_config["config"] = []
            config_temp = {}
            logger.info("{} found python file : {}".format("system$%util$%", file_name), __name__)
            if os.path.exists(file_name):
                path, file = Utils.get_path(file_name)
                sys.path.append(path)
                module_name = file.rsplit(".", 1)[0]

                config_temp["name"] = module_name
                config_temp["benchname"] = module_name

                module = importlib.import_module(module_name)

                class_list = {module_name: Utils.get_class_from_py(module)}
                if not class_list:
                    raise FailLoadCaseFromPyException

                case_dict = Utils.get_case_from_class(class_list)
                cases = Utils.load_dict_to_case(case_dict)
                return {
                    "config": [{
                        "name": module_name,
                        "benchname": module_name,
                        "module_name": module_name,
                        "module_path": path
                    }],
                    "testcases": cases
                }

        except Exception as e:
            logger.error(f"{file_name} {e} ", __name__)
            logger.error(f"{file_name} {traceback.format_exc()}", __name__)

    @staticmethod
    def http_request(request:Request, **kwargs) :
        '''
        http request
        :param request: Request
        :return: Response
        '''
        key = kwargs.get("key")

        header = request.header
        data = request.data
        method = request.method
        url = request.url

        response = Response()
        result = None
        if method == CASE_HTTP_API_METHOD_GET:
            result = HttpController.get(url, data, header, **kwargs)
        elif method == CASE_HTTP_API_METHOD_POST:
            result = HttpController.post(url, data, header)
        if method == CASE_HTTP_API_METHOD_PUT:
            result = HttpController.put(url, data, header)

        logger.info("{} {}".format(key, result), __name__)
        if isinstance(result, HTTPResponse):
            response.header = result.headers
            try:
                response.body = json.loads(result.read())
            except :
                response.body = result.read()
            response.code = result.code
        elif isinstance(result,urllib.request.URLError) or isinstance(result, urllib.request.HTTPError) or isinstance(result, urllib.request.HTTPHandler):
            response.body =  {}
            response.code = result.code if hasattr(result, "code") else None
            response.header = result.info() if hasattr(result, "info") else None

        logger.info(f"{key} {result}", __name__)
        return response

    @staticmethod
    def get_datetime_now():
        '''
        get datetime now to str
        :return: time now str
        '''
        current_time = time.time()
        local_time = time.localtime(current_time)
        time_temp = time.strftime("%Y-%m-%d %H:%M:%S", local_time)
        secs = (current_time - int(current_time)) * 1000
        timenow = "%s %03d" % (time_temp, secs)
        return timenow

    @staticmethod
    def get_date_result(begin: str, end: str)->int:
        '''
        get_date_result : get duration from begin to end
        :param begin: str
        :param end: str
        :return: int
        '''
        if begin is None or end is None:
            return 0
        try:
            d1, s1 = begin.rsplit(" ", 1)
            d2, s2 = end.rsplit(" ", 1)
            d1 = datetime.strptime(d1, "%Y-%m-%d %H:%M:%S")
            s1 = float(s1)
            d2 = datetime.strptime(d2, "%Y-%m-%d %H:%M:%S")
            s2 = float(s2)
            sec = (d2-d1).seconds
            if sec == 0:
                return (s2-s1)/1000
            else:
                return float(sec) + (s2-s1)/1000
        except Exception as e:
            return 0


    @staticmethod
    def get_time_str():
        '''
        get datetime now to str
        :return: time now str
        '''
        current_time = time.time()
        local_time = time.localtime(current_time)
        time_temp = time.strftime("%Y-%m-%d-%H.%M", local_time)
        return time_temp

    @staticmethod
    def jsontool(input):
        '''
        deal with json object
        :param input:
        :return:
        '''
        output = json.loads(input)
        return output

    @staticmethod
    def get_random_name():
        '''
        :return:
        '''
        return str(random.randint(100000, 999999))

    @staticmethod
    def get_platform():
        return platform.system()


class LoadFromConfig(object):

    @staticmethod
    def load_from_xlsx(file_name):
        if isinstance(file_name, str):
            inputs = Utils.get_rows_from_xlsx(file_name)
            return inputs

    @staticmethod
    def load_from_json(file_name):
        if isinstance(file_name, str):
            inputs = Utils.load_from_json(file_name)
            return inputs

    @staticmethod
    def load_from_yml(file_name):
        if isinstance(file_name, str):
            inputs = Utils.load_from_yml(file_name)
            return inputs

    @staticmethod
    def load_from_py(file_name):
        if isinstance(file_name, str):
            inputs = Utils.load_from_py(file_name)
            return inputs


class Signal(metaclass=SingletonType):
    def __init__(self):
        super().__init__()
        self.signal = True

    def change_status(self):
        self.signal = not self.signal


class SignalThread(Thread):
    def __init__(self, signal, signal_change_time=0.1):
        super().__init__()
        self.signal = signal
        self.time = signal_change_time
        self.daemon = True

    def run(self):
        while True:
            # logger.debug(f"signal of {self.signal.signal}", __name__)
            self.signal.change_status()
            time.sleep(self.time)

